"""
CySec Tools — CommandCenter Plugin
id: cysec  |  version: 1.0.0.9

Features:
  1. IVA Formatter — format IVA spreadsheets (manual + auto-watch Downloads)
  2. Quick Access  — one-click open X:\\scripts and Z:\\Fortinet\\FortiEMS
  3. Response Search — fast search across pre-written IVA response statements
"""
from __future__ import annotations

import os
import re
import sys
import subprocess
import threading
import datetime
from pathlib import Path
from typing import Optional

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QWidget, QStackedWidget, QTextEdit, QLineEdit, QCheckBox,
    QListWidget, QListWidgetItem, QSplitter, QScrollArea,
    QFrame, QFileDialog, QSizePolicy, QSpacerItem, QApplication,
    QAbstractItemView,
)
from PySide6.QtCore import Qt, QSize, QTimer, QUrl
from PySide6.QtGui import QFont, QDesktopServices, QColor

# ─── optional openpyxl ────────────────────────────────────────────────────────
try:
    import openpyxl                      # type: ignore
    _OPENPYXL_OK = True
except ImportError:
    openpyxl = None                      # type: ignore
    _OPENPYXL_OK = False

# ─── file/path constants ──────────────────────────────────────────────────────
_IVA_RE           = re.compile(r'^[A-Za-z0-9]+_IVA_\d+(\s.+)?\.xlsx$', re.IGNORECASE)
_FORMATTED_RE     = re.compile(r'\s+(?:COMPLETE|WIP)\.xlsx$', re.IGNORECASE)  # already formatted
_DUP_RE           = re.compile(r'\s*\(\d+\)\s*$')
_SCRIPTS_PATH = r"X:\scripts"
_FORTI_PATH   = r"Z:\Fortinet\FortiEMS"

# ─── response raw data ────────────────────────────────────────────────────────
_RAW: list[tuple[str, str]] = [
    # ── Vendor-Managed / Third-Party Devices ─────────────────────────────────
    ("Vendor-Managed / Third-Party Devices",
     "This device is vendor managed. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is vendor managed. It is recommended to reach out to Finastra to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a vendor managed device. It is recommended to reach out to Verifone to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "The use of HTTP is set up by design for Verifone functionality. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a vendor managed Cash Recycler Machine. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a vendor managed ATM/ITM. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a vendor managed Wireless Access Point (NetApp). It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a vendor managed server. It is recommended to reach out to DCI to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This is the vendor managed Core Director Access device. It is recommended to reach out to Jack Henry to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is managed by Fiserv. It is recommended to reach out to the vendor to have them apply the resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed phone system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed phone system. It is recommended to reach out to Shoretel to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed Zixgateway setup. It is recommended to reach out to ZixCorp to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This is a Zixgateway certificate. It is recommended to reach out to ZixCorp to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed Converge IVR system. It is recommended to reach out to Fiserv to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed camera system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This Wycom device is part of the vendor managed Check Printer setup. It is recommended to reach out to Jack Henry to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of Per Mar Security managed security system. It is recommended to reach out to Per Mar Security to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the vendor managed system. It is recommended to reach out to Rubrik to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a third-party audit scanner. It is recommended to reach out to Trace Security to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This certificate is for the MoveIT installation on the server. It is recommended to reach out to Ipswitch to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This is a LogMeIn certificate. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Managed / Third-Party Devices",
     "This is an HP iLO and is vendor managed. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Managed / Third-Party Devices",
     "This is an ONTAP System Manager and is vendor managed. It is recommended to reach out to NetApp to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is an IP camera system. It is recommended to reach out to Synology to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a Daktronics Digital Sign. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is an Entrust Datacard printer. It is recommended to reach out to Datacard to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is a Card printer. It is recommended to reach out to the debit card printer vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This EPSON TMNet printer is part of a vendor managed POS system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the Bank's DVR system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Managed / Third-Party Devices",
     "This device is part of the Bank's Camera system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Managed / Third-Party Devices",
     "The cipher is used for access to the Broadcom Management Service. It is recommended to reach out to Broadcom to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Managed / Third-Party Devices",
     "The SQL instance was set up on the device for Milestone XProtect, and modifying it could cause adverse effects. It is recommended to contact the vendor to apply the vulnerability resolutions."),
    ("Vendor-Managed / Third-Party Devices",
     "PostgreSQL is vendor managed software. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),

    # ── Bank-Managed / Navanta Lacks Access ──────────────────────────────────
    ("Bank-Managed / Navanta Lacks Access",
     "This device is Bank managed equipment that Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This NAS device is Bank managed equipment that Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This device is legacy Bank managed equipment that Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This server is being managed by Bank staff. Please contact Navanta at 877-778-7774 opt. 1 if you would like to proceed with applying the resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Bank-Managed / Navanta Lacks Access",
     "Navanta does not have access to the workstation to apply the resolution. Please reach out to Navanta at 877-778-7774 opt. 1 if you would like to proceed with applying the resolutions to this device. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Bank-Managed / Navanta Lacks Access",
     "This switch is managed by the Bank and Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This device is a legacy UPS, which Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This device is also inaccessible to hosts outside the internal network by means of both the financial institution's firewall policy and Network Address Translation (NAT), which further reduces any risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This Clickshare Configurator is legacy equipment that Navanta lacks the ability to manage. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "The Zoom client is managed in the VMware Horizon VDI management that Bank On IT has no access to, as it is bank managed. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "The Dell Wyse Management Suite software is Bank managed, so no changes were made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "The Manage Engine ADManager Plus software is Bank managed, so no changes were made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "The Manage Engine ServiceDesk Plus software is Bank managed, so no changes were made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "Tenable Nessus is Bank managed software. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This is a bank managed software that requires a license. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Bank-Managed / Navanta Lacks Access",
     "This site is set up by design for Bank managed software functionality. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── Vendor-Specific Software / Functionality by Design ───────────────────
    ("Vendor-Specific Software / Functionality by Design",
     "The site is set up by design for Kyocera functionality. It is recommended to reach out to Genesis to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of VMware Spring Framework is set up for Kyocera Gateway for Windows, which has been used with printer vendors to monitor equipment. It is recommended to reach out to your printer vendor to have them apply the vulnerability resolutions. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert. If the software is no longer in use, please call 1-877-778-7774 and submit a ticket to have it removed."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of Apache Tomcat is set up by design for Fiserv Store and Forward functionality. It is recommended to reach out to Fiserv to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of DB2 is set up for Store and Forward functionality. It is recommended to reach out to Fiserv to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of ASP .NET MVC 3 is set up for Abrigo BAM software functionality. It is recommended to reach out to Abrigo to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of HTTP is set up for CrushFTP Service functionality. It is recommended to reach out to the vendor to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "The use of FTP is set up by design for NBS functionality. It is recommended to reach out to NBS to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This is set up by design for NBS functionality. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "Java is installed on the server for Compliance Concierge. It is recommended to reach out to the vendor, FIPCO, to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "It is recommended to reach out to FIPCO to have them apply the vulnerability resolutions for the Oracle WebLogic Server site. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "Apache Log4j is in use for Wolters Kluwer software. It is recommended to reach out to Wolters Kluwer Financial Services to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This device is part of the vendor managed MEA Gateway. It is recommended to reach out to MEA Financial Enterprises to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This device is Intelli-M Supervisor/Infinias managed and has been set up by design for functionality. It is recommended to reach out to the vendor to have them apply the vulnerability resolution. Consequently, no changes have been made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal Bank network by means of both institution firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This device is the American Megatrends iKVM. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This device is the Verizon Wireless Network Extender for Business. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This service is set up for Spiceworks. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Specific Software / Functionality by Design",
     "This site is set up for Lansweeper software. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "Microsoft Access Database Engine is set up on the device for use with the OFAC Partner Program. Updating or modifying Microsoft Access Database Engine could disrupt the functionality of the device. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Vendor-Specific Software / Functionality by Design",
     "VMware Tools is set up on the device. Please contact Navanta if this software is no longer in use and can be uninstalled. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This software, Output Messenger, does not currently support the requested change. Navanta will continue to monitor for potential upgrades from the software developer. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Vendor-Specific Software / Functionality by Design",
     "This device is part of the Bank's Sharepoint Services. Consequently, no changes have been made at this time. To coordinate a quarterly update to these devices, please contact your Navanta Client Success Manager at 877-778-7774 opt. 1."),

    # ── ATMs ─────────────────────────────────────────────────────────────────
    ("ATMs",
     "This device is an ATM. It is vendor setup by design for functionality. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("ATMs",
     "This device is an ATM. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),

    # ── Printers ─────────────────────────────────────────────────────────────
    ("Printers",
     "This device is a printer. Unsupported changes to printers have been known to cause a loss of functionality including: Scan to Folder, Scan to Email, Printing to other Branches, and Printing Directly from Vendor Applications. Consequently, Navanta does not make changes to printer settings or firmware unless specifically requested by a client. This device is also inaccessible to hosts outside the internal network by means of both the Financial Institution's firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability. If you would like to work with Navanta to explore options for remediation of this finding, please contact your Navanta Strategic Advisor."),
    ("Printers",
     "This device is a printer with limited capabilities to minimize vulnerability risks. However, it has been observed that changing or disabling the SNMP Default Community String may affect the ability to print from financial vendor applications. Therefore, no modifications have been made to the default configuration settings at this time. Additionally, this device is inaccessible to hosts outside the internal network by means of both the financial institution's firewall policy and Network Address Translation (NAT), which further reduces the risk introduced by this vulnerability."),
    ("Printers",
     "This device is a printer. The resolution for this vulnerability requires a firmware update. Firmware updates are known to possibly cause several issues including the loss of printer functionality. Consequently, no changes have been made at this time. This device is also inaccessible to hosts outside the internal network by means of both the financial institution's firewall policy and Network Address Translation (NAT), which further reduces any risk introduced by this vulnerability. If you would like to work with Navanta to explore options for remediation of this finding, please contact your Navanta Strategic Advisor."),
    ("Printers",
     "This is an FTP site set up by design for the Scan to Folder functionality, and no changes were made as it would cause issues with the FTP site. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Printers",
     "The referenced Windows Update (KB5005010) has been applied to this device. Additional settings by Microsoft were introduced alongside this update to allow restrictions on installing or updating device drivers to require Administrator rights. These settings may change the day-to-day operation of printers at the institution and may not be compatible with all types of printers, such as certain configurations of teller receipt printers. Please contact Navanta if you would like to proceed with exploring these additional requirements, 877-778-7774, opt 1."),
    ("Printers",
     "To update the default credentials of the Polycom SoundStation, a factory reset would be needed. No changes were made at this time. Please contact Navanta if you would like to proceed with updating the default credentials, at 877-778-7774 opt. 1."),

    # ── Security / Alarm Systems ──────────────────────────────────────────────
    ("Security / Alarm Systems",
     "This device is part of the Bank's alarm system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Security / Alarm Systems",
     "This device is part of the Bank's security system. It is recommended to reach out to the vendor to have them apply the vulnerability resolutions. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── Winterms / Thin Clients ───────────────────────────────────────────────
    ("Winterms / Thin Clients",
     "This device is a winterm, limiting its capabilities of being updated and compliant. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Winterms / Thin Clients",
     "This device is a winterm. Changing protocols and ciphers on winterms may cause functionality issues. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Winterms / Thin Clients",
     "This device is a winterm. Winterms lack the resources to install Microsoft patches. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Winterms / Thin Clients",
     "The winterms are in the process of being replaced, so no changes were made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── End-of-Life Operating Systems / Software ──────────────────────────────
    ("End-of-Life Operating Systems / Software",
     "This device is running Windows XP, limiting its capabilities of being updated and compliant. It is recommended to remove this device from the network. Please contact Navanta if you would like to proceed with replacing or removing the device, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Windows 7, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported operating system on the device. Please contact Navanta if you would like to proceed with replacing or removing the device, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "Windows 10 has reached End of Life as of October 2025. It is recommended to upgrade this device to Windows 11 to resolve this finding. If it is not compatible with Windows 11, it is recommended to replace this device. This device is also inaccessible to hosts outside the internal network by means of both the financial institution's firewall policy and Network Address Translation (NAT), which further reduces any risk introduced by this vulnerability. If you would like to work with Navanta to explore options for remediation of this finding, please contact your Navanta Strategic Advisor."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Windows Server 2003 R2, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported operating system on the device. Please contact Navanta if you would like to proceed with replacing or removing the device, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Windows Server 2012 R2, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported operating system on the device. Please contact Navanta if you would like to proceed with replacing or removing the device, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Microsoft SQL 2008, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported version of Microsoft SQL on the device. Please contact Navanta if you would like to proceed with updating the software, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Microsoft SQL 2012, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported version of Microsoft SQL on the device. Please contact Navanta if you would like to proceed with updating the software, 877-778-7774, opt 1."),
    ("End-of-Life Operating Systems / Software",
     "This device is running Microsoft SQL 2014 SP3, which is end of life. This limits its capabilities of being updated and compliant. It is recommended to install a supported version of Microsoft SQL on the device. Please contact Navanta if you would like to proceed with updating the software, 877-778-7774, opt 1."),

    # ── Certificates (Self-Signed) ────────────────────────────────────────────
    ("Certificates (Self-Signed)",
     "This device sits within the network and is only accessed from other devices within the network. It was decided that a self-signed certificate from the manufacturer was appropriate for this device and type of connection, and that the benefits that would come with a third-party certificate would not outweigh the costs."),
    ("Certificates (Self-Signed)",
     "The certificate on the server was created and authenticated by SQL. It was decided that a self-signed certificate was appropriate for this device and type of connection, and that the benefits that would come with a third-party certificate would not outweigh the costs. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Certificates (Self-Signed)",
     "This self-signed certificate is valid and is installed for FortiClient on this device. We will continue to monitor for potential updates released by FortiNet that enable this functionality in the future. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── FortiClient / FortiEMS ────────────────────────────────────────────────
    ("FortiClient / FortiEMS",
     "Forticlient EMS is currently running the latest approved version. Navanta is reviewing this alert and will apply the resolution during the normal patching process."),
    ("FortiClient / FortiEMS",
     "Forticlient version 7.0.14.585 is in use to maintain connection with the EMS server. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("FortiClient / FortiEMS",
     "The vulnerable components of FortiClient TunnelVision are not in use for this device. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("FortiClient / FortiEMS",
     "The vulnerable components of FortiEMS are not in use for this device. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── TLS / SSL / Ciphers / Protocols ──────────────────────────────────────
    ("TLS / SSL / Ciphers / Protocols",
     "Applied secure SCHANNEL baseline configuration to the device, including the disabling of TLS 1.0 and TLS 1.1."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of the Diffie-Hellman Key Exchange group was disabled to resolve the alert."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of weak ciphers and protocols has been disabled on this device."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of the TLSv1 protocol is required for the software running on the device. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of the RC4 ciphers is required for the software running on the device. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of the Triple-DES ciphers is required for the software running on the device. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("TLS / SSL / Ciphers / Protocols",
     "CBC mode ciphers, RSA key exchanges, and the SHA-1 algorithm are supported to maintain backward compatibility."),
    ("TLS / SSL / Ciphers / Protocols",
     "Disabling the use of TLSv1 and TLSv1.1 protocols will be tested at Navanta via created projects to determine that disabling the protocols will not negatively impact the institution's business needs."),
    ("TLS / SSL / Ciphers / Protocols",
     "Testing to disable TLS 1.0 and 1.1 is currently underway within Navanta and will be deployed once confirmed it will not cause any disruptions of service. If you would like to work with Navanta to explore options for remediation of this finding, please contact your Navanta Strategic Advisor."),
    ("TLS / SSL / Ciphers / Protocols",
     "Disabled the reported weak KEX algorithm(s) — 1024-bit MODP group. The vendor DCI uses FTP to upload work. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("TLS / SSL / Ciphers / Protocols",
     "In ticket 1615895, DCI no longer needed the weak KEX algorithms, and they were removed from the server. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("TLS / SSL / Ciphers / Protocols",
     "The use of Main Mode authentication was set on the router to resolve the IKE Aggressive Mode alert."),

    # ── Updates / Patches Applied ─────────────────────────────────────────────
    ("Updates / Patches Applied", "A firmware update was applied to resolve this alert."),
    ("Updates / Patches Applied", "The Microsoft update was applied to the device."),
    ("Updates / Patches Applied",
     "The Microsoft update was applied to the device, and a group policy to set the registry key value is currently being tested before being implemented bank-wide."),
    ("Updates / Patches Applied", "The certificate was reissued on the device to resolve the alert."),
    ("Updates / Patches Applied", "The ActiveX killbit was set on the device to resolve the alert."),
    ("Updates / Patches Applied", "The service was disabled via registry edits to resolve the alert."),
    ("Updates / Patches Applied", "The default credential was updated to resolve the alert."),
    ("Updates / Patches Applied", "The MSXML4.dll files were removed from the device to resolve the alert."),
    ("Updates / Patches Applied",
     "MSXML Parser and XML Core services have been updated and the old version uninstalled to resolve this issue."),
    ("Updates / Patches Applied", "Removed the log4j-1.2.14.jar file to clear the alert."),
    ("Updates / Patches Applied",
     "Java was updated to version 8 update 201 and the older version was removed from the device to resolve the alert."),
    ("Updates / Patches Applied",
     "The quotes were added to the service paths on the device to resolve the alerts."),
    ("Updates / Patches Applied", "Added secure quotation to the registry paths."),
    ("Updates / Patches Applied",
     "A newer version of .NET Core was updated on the device and the older version was removed to resolve the alert."),
    ("Updates / Patches Applied", "Microsoft Office Click-to-Run was updated on the device."),
    ("Updates / Patches Applied",
     "Mozilla Firefox ESR was updated on the device to version 78.7.1 for DCI compatibility."),
    ("Updates / Patches Applied",
     "Microsoft Windows Snipping Tool has been updated on this device to resolve this alert."),
    ("Updates / Patches Applied", "Python was updated to the most recent version."),
    ("Updates / Patches Applied", "Adobe Media Encoder was updated to the most recent version."),
    ("Updates / Patches Applied", "Microsoft Web Media Extensions have been removed from the system."),
    ("Updates / Patches Applied",
     "OpenSSL was part of McAfee Connect, which has been uninstalled from the device to resolve the alert."),
    ("Updates / Patches Applied", "The application has been updated to resolve the alert."),
    ("Updates / Patches Applied", "The application has been removed to resolve the alert."),
    ("Updates / Patches Applied", "The updated device configuration was applied to resolve the alert."),
    ("Updates / Patches Applied", "The firewall configuration was updated to resolve this alert."),
    ("Updates / Patches Applied", "The device was decommissioned on ticket 1468635 after the scan was completed."),
    ("Updates / Patches Applied",
     "The device was onboarded in ticket #1692443, and the Microsoft updates were applied to resolve the alert."),
    ("Updates / Patches Applied",
     "Although OneDrive is not in use on this device, a change to firewall policy has been made to allow future updates to the software. (Ticket #1511306)"),
    ("Updates / Patches Applied",
     "This device has been updated to Windows 11 and the latest version of Office 2016."),
    ("Updates / Patches Applied", "This device has been removed from the network."),

    # ── Updates In Progress / Under Testing ───────────────────────────────────
    ("Updates In Progress / Under Testing",
     "This update is currently undergoing internal testing at Navanta and will be applied once the update has been confirmed to not negatively impact the Bank's business needs. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Updates In Progress / Under Testing",
     "This patch has been released within the past 30 days of the scan and will be installed on the device once it is made available to the Bank by the Navanta Management Agent."),
    ("Updates In Progress / Under Testing",
     "This patch has been released within the past 60 days and will be installed on the device as part of the standard patch management cycle."),
    ("Updates In Progress / Under Testing",
     "The referenced update did not successfully pass through testing due to issues it caused with printing. Navanta has deployed other mitigations to prevent exploitation of this vulnerability."),
    ("Updates In Progress / Under Testing",
     "A group policy is currently being deployed to identify if setting the registry key value can be made without disrupting the functionality of the device. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Updates In Progress / Under Testing",
     "This update is currently undergoing Navanta testing and will be deployed once it is confirmed to have no negative impact."),
    ("Updates In Progress / Under Testing",
     "Port 6060 is for ReadySite. This update is currently undergoing internal testing at Navanta and will be applied once the update has been confirmed to not negatively impact the Bank's business needs. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),

    # ── Being Researched / Ticket In Progress ─────────────────────────────────
    ("Being Researched / Ticket In Progress",
     "This finding was still being researched at the time the report was published. An updated response to this finding will be in the next Quarterly IVA. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Being Researched / Ticket In Progress",
     "The resolution is being researched and will be applied on ticket number <case#>."),

    # ── No Update / Fix Available ─────────────────────────────────────────────
    ("No Update / Fix Available",
     "There is currently no update available at this time. Navanta will continue to monitor for potential updates released by the manufacturer. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "This device does not support the requested change at this time. Navanta will continue to monitor for potential updates released by the manufacturer that enable this functionality in the future. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "This device does not currently support the requested change. Navanta will continue to monitor for potential upgrades from the device manufacturer. Consequently, no changes have been made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability. It is recommended to replace the device."),
    ("No Update / Fix Available",
     "HP Device Manager is set up for maintaining backwards compatibility with legacy devices. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "This is the Tripp Lite Power Alert UPS device. Unsupported changes to the APC UPS have been known to cause a loss of functionality, so no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "No changes were made to the software on the device to maintain compatibility between the iDRAC components. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "Zoom has not released an update at this time to resolve the screen sharing vulnerability in Windows. It is recommended to use best practices for screen sharing while using Zoom Client, such as presenting from a clear desktop background or selectively sharing one screen, to minimize risk. Navanta will continue to monitor for potential updates released by the manufacturer. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "Zoom Outlook Plugin is on the most stable version available. Navanta will continue to monitor for potential updates released by the manufacturer. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "This device does not currently support the use of SSH as a management protocol. Consequently, no changes have been made at this time. This device is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("No Update / Fix Available",
     "This device is at end of life and is no longer supported by the manufacturer. Access to the website is denied over port 80; no request for credentials comes up."),

    # ── Approval Required / Licensing Cost ───────────────────────────────────
    ("Approval Required / Licensing Cost",
     "It is recommended to install a newer version of Adobe Acrobat on this station. There is a licensing cost to perform this update. If you would like to review updating Adobe Acrobat at this time, please contact Navanta at 877-778-7774, opt. 1. #Approval"),
    ("Approval Required / Licensing Cost",
     "Adobe Creative Cloud requires login credentials to be updated. Please contact Navanta if you would like to proceed with updating or removing the software, 877-778-7774, opt 1. #Approval"),
    ("Approval Required / Licensing Cost",
     "It is recommended to install a newer version of Java on this device. There is a licensing cost to perform this update. If you would like to review updating Java at this time, please contact Navanta at 877-778-7774, opt. 1. #Approval"),
    ("Approval Required / Licensing Cost",
     "It is recommended to install a newer version of Foxit PhantomPDF on this station. There is a licensing cost to perform this update. If you would like to review updating Foxit PhantomPDF at this time, please contact Navanta at 877-778-7774, opt. 1. #Approval"),
    ("Approval Required / Licensing Cost",
     "It is recommended to install a newer version of %Software% on this device. There is a licensing cost to perform this update. If you would like to review updating %Software% at this time, please contact your Navanta Strategic Advisor. Consequently, no changes have been made at this time. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this vulnerability."),
    ("Approval Required / Licensing Cost",
     "It is recommended to update Skype to a newer version. However, Skype requires login credentials to be updated. Please contact Navanta if you would like to proceed with updating or removing the software. #Approval"),
    ("Approval Required / Licensing Cost",
     "Though it is recommended to update Mozilla Firefox to a later version, the Bank has requested to not do so for risk of impacting applications running on the device. If you would like for Navanta to update Mozilla Firefox on the device, please contact Navanta Support at 877-778-7774, opt 1."),
    ("Approval Required / Licensing Cost",
     "Though it is recommended to update Google Chrome to a later version, the Bank has requested to not do so for risk of impacting applications running on the device. If you would like for Navanta to update Google Chrome on the device, please contact Navanta Support at 877-778-7774, opt 1."),

    # ── Network / Gateway / DMZ ───────────────────────────────────────────────
    ("Network / Gateway / DMZ",
     "This network is a DMZ, and the configuration is set up by design. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Network / Gateway / DMZ",
     "This is a gateway device and represents the first hop outside of the network. ICMP services are an intentional part of its configuration as part of Navanta's services as the Bank's ISP."),
    ("Network / Gateway / DMZ",
     "This is a gateway device and represents the first hop outside of the network. External access is an intentional part of its configuration as part of Navanta's services as the Bank's ISP."),
    ("Network / Gateway / DMZ",
     "This login page is necessary for the Bank's VPN configuration. The use of strong and complex passwords is required, failed logins are monitored, and account lockout policies are implemented."),
    ("Network / Gateway / DMZ",
     "This login page is necessary for the Bank's Outlook Web Access. The use of strong and complex passwords is required, failed logins are monitored, and account lockout policies are implemented."),
    ("Network / Gateway / DMZ",
     "ICMP is enabled for internal network communication to provide system and network updates. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Network / Gateway / DMZ",
     "The Timestamp reply is intentionally enabled on this Navanta equipment for internal network diagnostics and is not exposed to external networks. The information disclosed (system time) is not sensitive and poses no security risk. Therefore, no changes to the equipment's configuration have been made."),
    ("Network / Gateway / DMZ",
     "This service is enabled as part of the Bank's email functionality. SMTP services are an intentional part of its configuration as part of Navanta's services as the Bank's ISP. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Network / Gateway / DMZ",
     "The Cisco Smart Install Protocol is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Network / Gateway / DMZ",
     "The VNC service is running locally on the internal network. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert. Consequently, no changes have been made at this time."),

    # ── DNS / DNSSEC ──────────────────────────────────────────────────────────
    ("DNS / DNSSEC",
     "To save system and network resources, DNS caching is used. DNS caching should only be disabled on DNS servers that are accessible to the public. Because of the firewall policy and Network Address Translation (NAT), it is inaccessible to hosts outside the internal network, lowering the risk posed by this alert."),
    ("DNS / DNSSEC",
     "DNSSEC is implemented for .bank domains, which is provided at an additional cost. To discuss implementing this change, please contact Navanta at 1-877-778-7774, Option 1: Approvals."),

    # ── Informational / No Action Required ───────────────────────────────────
    ("Informational / No Action Required",
     "This finding is informational. The information obtained poses no security risk."),
    ("Informational / No Action Required",
     "The Microsoft Exchange server is configured to provide requested information per Microsoft's best practice and does not disclose any information not necessary for the functionality of the server. Currently, outside of disabling NTLM authentication over HTTP, there is no method to mitigate the leaking of such information under Microsoft IIS — all versions are affected by design. No action is required at this time."),
    ("Informational / No Action Required",
     "The scan identified standard Windows communication services running on devices within the private internal network. These services are a normal part of Microsoft Windows and are used for everyday system operations such as device management and internal network communication. Because the scan was conducted on a private local network, the presence of these services is expected and does not indicate malicious activity. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Informational / No Action Required",
     "This device automatically redirects its HTTP traffic to HTTPS, reducing any risk associated with this alert."),
    ("Informational / No Action Required / HID",
     "No solution has been provided by the vendor; access to USB drives is restricted by domain level policies on all workstations. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Informational / No Action Required",
     "The vendor has supplied no solution to this alert. Navanta restricts folder access through the means of group policy and access controls, reducing the risk created by this alert. Consequently, no changes have been made at this time. This device is still inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Informational / No Action Required",
     "Navanta AntiVirus is in use on the device. Consequently, no changes have been made at this time. This service is inaccessible to hosts outside the internal Bank network by means of both Bank firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Informational / No Action Required",
     "Navanta has verified this device is not vulnerable in its current configuration."),
    ("Informational / No Action Required",
     "The referenced CVE was later revised by the vendor. The identified vulnerability only applies to other versions of Microsoft Web Deploy, which are not present on this system. The version installed on this device is not affected."),

    # ── Offline / Decommissioned ──────────────────────────────────────────────
    ("Offline / Decommissioned",
     "This device was offline during various attempts to resolve alerts. Navanta runs quarterly Vulnerability Assessments for the client, so if this device is in use again or turned back on, it should be detected in a future scan."),

    # ── Device Onboarding ─────────────────────────────────────────────────────
    ("Device Onboarding",
     "The onboarding process for this device has not been completed. Please contact your Navanta Strategic Advisor to discuss onboarding this device."),

    # ── Sister Bank Alerts ────────────────────────────────────────────────────
    ("Sister Bank Alerts",
     "This alert is being addressed in the Vulnerability Assessment for Bank of Cordell's Sister Bank, Bank of Hydro."),
    ("Sister Bank Alerts",
     "This alert is being addressed in the Vulnerability Assessment for Chillicothe State Bank's Sister Bank, the State Bank of Richmond."),

    # ── Secondary IP Address ──────────────────────────────────────────────────
    ("Secondary IP Address",
     "This is a secondary IP address for the device located at xxx.xxx.xxx.xxx. The secondary IP address is necessary for device functionality. Please see the corresponding alert at the primary IP address for a response to this alert."),

    # ── Miscellaneous ─────────────────────────────────────────────────────────
    ("Miscellaneous",
     "The ReadySite Backup vendor has been contacted about updating their software to a more recent version of PHP. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Miscellaneous",
     "Adobe Flash is a part of Windows 8.1/10, including Windows Server 2008, 2012, and 2016. This will be updated after the next Windows Microsoft update."),
    ("Miscellaneous",
     "BLANK is setup for BLANK functionality. It is recommended to reach out to the vendor to have them apply the vulnerability resolution. To provide additional security to the device, it is inaccessible to hosts outside the internal network by means of both the client firewall policy and Network Address Translation (NAT), which reduces the risk introduced by this alert."),
    ("Miscellaneous",
     "It is recommended to upgrade this device to Windows 11 to resolve this finding. (Microsoft Windows Elevation of Privilege Vulnerability — KB5042320)"),
]

# ─── auto-tag keyword map ────────────────────────────────────────────────────
_KW_TAGS: dict[str, list[str]] = {
    "nat":            ["nat ", "network address translation"],
    "firewall":       ["firewall"],
    "vendor":         ["vendor managed", "vendor", "third-party"],
    "bank-managed":   ["bank managed", "bank staff"],
    "navanta":        ["navanta"],
    "no-changes":     ["no changes have been made"],
    "resolved":       ["was applied", "has been updated", "has been removed",
                       "was updated", "applied to resolve", "updated to resolve",
                       "removed to resolve", "cleared the alert"],
    "in-progress":    ["undergoing", "being researched", "ticket"],
    "approval":       ["licensing cost", "#approval", "contact navanta"],
    "informational":  ["informational", "no action required", "no security risk"],
    "atm":            ["atm", "itm"],
    "printer":        ["printer", "scan to folder"],
    "server":         ["server"],
    "switch":         ["switch"],
    "ups":            ["ups"],
    "camera":         ["camera", "dvr"],
    "winterm":        ["winterm"],
    "gateway":        ["gateway", "dmz"],
    "tls-ssl":        ["tls", "ssl", "protocol", "cipher", "schannel"],
    "certificate":    ["certificate", "self-signed"],
    "patch":          ["patch", "microsoft update", "firmware update"],
    "eol":            ["end of life", "end-of-life"],
    "windows":        ["windows"],
    "fiserv":         ["fiserv"],
    "jack-henry":     ["jack henry"],
    "verifone":       ["verifone"],
    "finastra":       ["finastra"],
    "forticlient":    ["forticlient", "fortiems", "fortinet"],
    "zixcorp":        ["zixcorp", "zixgateway"],
    "microsoft":      ["microsoft"],
    "cisco":          ["cisco"],
    "netapp":         ["netapp"],
    "java":           ["java"],
    "adobe":          ["adobe"],
    "zoom":           ["zoom"],
    "vmware":         ["vmware"],
    "kyocera":        ["kyocera"],
    "synology":       ["synology"],
    "hp":             ["hp ilo", "hp device"],
    "rubrik":         ["rubrik"],
    "dci":            ["dci"],
    "dns":            ["dnssec", " dns "],
    "http-ftp":       ["http", "ftp"],
    "smtp":           ["smtp"],
    "icmp":           ["icmp"],
    "vnc":            ["vnc"],
    "vpn":            ["vpn"],
}


def _build_tags(category: str, text: str) -> list[str]:
    tags: set[str] = set()
    cat_tag = re.sub(r'[^a-z0-9]+', '-', category.lower()).strip('-')
    tags.add(cat_tag)
    tl = text.lower()
    for tag, kws in _KW_TAGS.items():
        if any(kw in tl for kw in kws):
            tags.add(tag)
    return sorted(tags)


_RESPONSES: list[dict] = [
    {"id": i + 1, "category": cat, "text": txt, "tags": _build_tags(cat, txt)}
    for i, (cat, txt) in enumerate(_RAW)
]
_ALL_CATEGORIES: list[str] = sorted({r["category"] for r in _RESPONSES})

# ─── module globals ───────────────────────────────────────────────────────────
_api: Optional[object]              = None
_btn                                = None
_dlg: Optional["CySecDialog"]       = None
_watcher_timer                      = None
_watcher_enabled: bool              = False
_watcher_start_dt: Optional[datetime.datetime] = None
_DOWNLOADS          = Path.home() / "Downloads"
_DEFAULT_OUTPUT_DIR = Path.home() / "Documents" / "IVA Reports" / "WIP"


def _get_output_dir() -> Path:
    """Return the configured output directory, falling back to the default."""
    saved = _api.settings.value("output_dir", "") if _api else ""
    p = Path(saved) if saved else _DEFAULT_OUTPUT_DIR
    p.mkdir(parents=True, exist_ok=True)
    return p


# ═══════════════════════════════════════════════════════════════════════════════
# Plugin entry points
# ═══════════════════════════════════════════════════════════════════════════════

def activate(api):
    global _api, _btn, _watcher_enabled, _watcher_start_dt, _watcher_timer
    _api = api
    _btn = api.ui.add_footer_button("CySec", _open)

    if api.settings.value("auto_format_enabled", False):
        iso = api.settings.value("watcher_start_iso", "")
        try:
            _watcher_start_dt = (datetime.datetime.fromisoformat(iso)
                                 if iso else datetime.datetime.now())
        except (ValueError, TypeError):
            _watcher_start_dt = datetime.datetime.now()
        _watcher_enabled = True
        _watcher_timer = api.timers.create(5000, _poll_downloads)
        api.log("CySec: downloads watcher restored from settings")


def deactivate():
    global _api, _btn, _dlg
    try:
        if _btn is not None:
            _api.ui.remove_footer_button(_btn)
            _btn = None
    except Exception:
        pass
    try:
        if _dlg is not None and _dlg.isVisible():
            _dlg.close()
        _dlg = None
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# Dialog opener
# ═══════════════════════════════════════════════════════════════════════════════

def _open():
    global _dlg
    try:
        if _dlg is not None and _dlg.isVisible():
            _dlg.raise_()
            _dlg.activateWindow()
            return
        _dlg = CySecDialog(_api.ui.main_window)
        _dlg.show()
    except Exception as e:
        _api.log(f"CySec _open error: {e}")
        _api.toast("Could not open CySec Tools.", "error")


# ═══════════════════════════════════════════════════════════════════════════════
# Watcher helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _set_watcher(enabled: bool):
    global _watcher_enabled, _watcher_start_dt, _watcher_timer
    _watcher_enabled = enabled
    if enabled:
        _watcher_start_dt = datetime.datetime.now()
        _api.settings.set("auto_format_enabled", True)
        _api.settings.set("watcher_start_iso", _watcher_start_dt.isoformat())
        if _watcher_timer is None:
            _watcher_timer = _api.timers.create(5000, _poll_downloads)
        _api.log("CySec: watcher enabled")
    else:
        _watcher_enabled = False
        _watcher_start_dt = None
        if _watcher_timer is not None:
            _api.timers.cancel(_watcher_timer)
            _watcher_timer = None
        _api.settings.set("auto_format_enabled", False)
        _api.settings.remove("watcher_start_iso")
        _api.log("CySec: watcher disabled")


def _poll_downloads():
    if not _watcher_enabled:
        return
    try:
        today = datetime.date.today()
        for f in _DOWNLOADS.glob("*.xlsx"):
            # Only consider files modified today
            try:
                if datetime.date.fromtimestamp(f.stat().st_mtime) != today:
                    continue
            except OSError:
                continue
            # Must match IVA filename pattern
            if not _IVA_RE.match(f.name):
                continue
            # Skip output files the formatter already produced
            if _FORMATTED_RE.search(f.name):
                continue
            # Try to format; if file is still being written, silently retry next poll
            _do_auto_format(f)
    except Exception as e:
        _api.log(f"CySec watcher poll error: {e}")


def _do_auto_format(src: Path):
    try:
        result = _format_iva_file(src)
        if result is None:
            # Already formatted — nothing to do (shouldn't reach here via auto,
            # since formatted files are moved out of Downloads)
            _api.log(f"CySec: skipped (already formatted): {src.name}")
            return
        new_name, status, dest = result
        _api.toast(f"IVA Auto-Formatted \u2192 {new_name}  [{status}]", "success")
        _api.log(f"CySec auto-formatted: {dest}")
        if _dlg and _dlg.isVisible():
            _dlg.iva_tab.append_log(f"[Auto] Formatted: {new_name}  [{status}]")
    except Exception as e:
        # File is likely still downloading — retry next poll
        _api.log(f"CySec auto-format retry pending: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# IVA formatting logic
# ═══════════════════════════════════════════════════════════════════════════════

def _clean_stem(name: str) -> str:
    stem = Path(name).stem
    return _DUP_RE.sub("", stem).strip()


def _is_green_cell(cell) -> bool:
    try:
        fill = cell.fill
        if fill.patternType == "solid":
            rgb = fill.fgColor.rgb
            return rgb.upper() in ("FF90EE90", "0090EE90", "90EE90")
        return False
    except (ValueError, AttributeError, TypeError):
        return False


def _check_complete(ws) -> bool:
    has_data = False
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=6)
        if cell.value is None or str(cell.value).strip() == "":
            continue
        has_data = True
        if not _is_green_cell(cell):
            return False
    return has_data


def _apply_iva_formatting(ws):
    from openpyxl.styles import Alignment

    # ── Row heights ───────────────────────────────────────────────────────────
    # Set an explicit height on every row (matches reference format).
    # Using per-row heights rather than sheet defaultRowHeight because Excel
    # renders them identically but per-row is what the reference sheets use.
    ws.sheet_format.defaultRowHeight = 15
    ws.sheet_format.customHeight = None
    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 50.1

    # ── Column widths ─────────────────────────────────────────────────────────
    # Set every used column to width 16.
    from openpyxl.utils import get_column_letter as _gcl
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[_gcl(col_idx)].width = 16

    # ── Cell alignment & empty-cell fill ─────────────────────────────────────
    # Reference format: wrap_text=True + vertical=top on every cell.
    # Header row (row 1) additionally gets horizontal=center.
    # Empty cells get value="" so Excel treats them as occupied (no overflow).
    for row in ws.iter_rows():
        is_header = (row[0].row == 1)
        for cell in row:
            if cell.value is None:
                cell.value = ""
            existing = cell.alignment
            h = existing.horizontal if existing and existing.horizontal else None
            if is_header:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="top",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal=h,
                    vertical="top",
                    wrap_text=True,
                )

    # ── Sheet-level settings ──────────────────────────────────────────────────
    # Auto-filter spans the actual used columns, not a hard-coded A1:M1.
    last_col = _gcl(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}1"
    # Do NOT set freeze_panes — reference sheets do not use it.
    ws.freeze_panes = None


def _is_already_formatted(ws) -> bool:
    """Return True if this worksheet already has IVA formatting applied.

    Checks the structural markers we set:
      • autofilter is set
      • row 1 has an explicit height of ~50
      • every data cell has wrap_text=True
    All must match to be considered formatted.
    """
    if not ws.auto_filter.ref:
        return False
    rd1 = ws.row_dimensions.get(1)
    if not rd1 or not rd1.height or rd1.height < 40:
        return False
    # If any cell lacks wrap_text the file needs re-formatting
    for row in ws.iter_rows():
        for cell in row:
            if not (cell.alignment and cell.alignment.wrap_text):
                return False
    return True


def _format_iva_file(src: Path) -> Optional[tuple[str, str, Path]]:
    """Format an IVA xlsx; returns (new_filename, status, dest_path), None if
    already formatted, or raises on error."""
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl is not installed.")
    wb = openpyxl.load_workbook(str(src))
    ws = wb.active
    if _is_already_formatted(ws):
        return None  # nothing to do
    sheet_name = ws.title
    status = "COMPLETE" if _check_complete(ws) else "WIP"
    _apply_iva_formatting(ws)
    clean = _clean_stem(src.name)
    base_name = f"{clean} {sheet_name} {status}.xlsx"
    out_dir = _get_output_dir()
    dest = out_dir / base_name
    if dest.exists():
        stem = dest.stem
        n = 1
        while (out_dir / f"{stem} ({n}).xlsx").exists():
            n += 1
        dest = out_dir / f"{stem} ({n}).xlsx"
    new_name = dest.name
    wb.save(str(dest))
    # Remove source from Downloads now that it has been moved to output dir
    try:
        src.unlink()
    except OSError:
        pass
    return (new_name, status, dest)


def _install_openpyxl():
    def _run():
        try:
            # sys._base_executable is the real .exe path; avoids the Windows
            # Store App-Execution-Alias reparse point that sys.executable may
            # point to, which causes subprocess to re-launch the app instead of
            # running pip.
            exe = (getattr(sys, "_base_executable", None) or sys.executable)
            _api.log(f"CySec: installing openpyxl via {exe}")
            r = subprocess.run(
                [exe, "-m", "pip", "install", "openpyxl"],
                capture_output=True,
                text=True,
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
            if r.returncode == 0:
                _api.toast(
                    "openpyxl installed — restart CommandCenter to enable IVA Formatter.",
                    "success")
            else:
                _api.toast("openpyxl install failed — check logs.", "error")
                _api.log(f"CySec pip stderr: {r.stderr[:400]}")
        except Exception as e:
            _api.log(f"CySec install_openpyxl error: {e}")
    threading.Thread(target=_run, daemon=True).start()
    _api.toast("Installing openpyxl…", "info")


# ═══════════════════════════════════════════════════════════════════════════════
# Shared style helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _sep(parent=None) -> QFrame:
    f = QFrame(parent)
    f.setFrameShape(QFrame.Shape.HLine)
    f.setObjectName("cs_sep")
    return f


# ═══════════════════════════════════════════════════════════════════════════════
# Main Dialog
# ═══════════════════════════════════════════════════════════════════════════════

class CySecDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("CySec Tools")
        self.setMinimumSize(960, 660)
        self.resize(1020, 700)
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint
        )
        self.iva_tab: Optional[IVAFormatterTab]     = None
        self.qa_tab:  Optional[QuickAccessTab]      = None
        self.rs_tab:  Optional[ResponseSearchTab]   = None
        self._tab_btns: list[QPushButton]           = []
        self._stack: Optional[QStackedWidget]       = None
        self._build_ui()
        _api.theme.register(self._apply_theme)
        self._apply_theme()

    def closeEvent(self, event):
        # Unregister the theme callback so the dialog can be GC'd
        try:
            _api.theme.unregister(self._apply_theme)
        except Exception:
            pass
        super().closeEvent(event)

    # ── build ─────────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # header
        hdr = QFrame()
        hdr.setObjectName("cs_hdr")
        hdr.setFixedHeight(52)
        hlay = QHBoxLayout(hdr)
        hlay.setContentsMargins(18, 0, 14, 0)
        lbl_icon  = QLabel("🔒")
        lbl_icon.setFont(QFont("Segoe UI Emoji", 15))
        lbl_title = QLabel("CySec Tools")
        lbl_title.setObjectName("cs_title_lbl")
        lbl_title.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        hlay.addWidget(lbl_icon)
        hlay.addSpacing(8)
        hlay.addWidget(lbl_title)
        hlay.addStretch()
        root.addWidget(hdr)

        # tab bar
        tbar = QFrame()
        tbar.setObjectName("cs_tbar")
        tbar.setFixedHeight(44)
        tlay = QHBoxLayout(tbar)
        tlay.setContentsMargins(14, 4, 14, 4)
        tlay.setSpacing(6)

        self._stack    = QStackedWidget()
        self.iva_tab   = IVAFormatterTab()
        self.qa_tab    = QuickAccessTab()
        self.rs_tab    = ResponseSearchTab()
        self._stack.addWidget(self.iva_tab)
        self._stack.addWidget(self.qa_tab)
        self._stack.addWidget(self.rs_tab)

        for idx, label in enumerate(("📊  IVA Formatter", "📁  Quick Access", "🔍  Response Search")):
            btn = QPushButton(label)
            btn.setObjectName(f"cs_tab_{idx}")
            btn.setCheckable(True)
            btn.setFixedHeight(32)
            btn.setFont(QFont("Segoe UI", 10))
            btn.clicked.connect(lambda _checked, i=idx: self._switch(i))
            tlay.addWidget(btn)
            self._tab_btns.append(btn)
        tlay.addStretch()

        root.addWidget(tbar)
        root.addWidget(_sep())
        root.addWidget(self._stack)

        self._switch(0)

    def _switch(self, idx: int):
        self._stack.setCurrentIndex(idx)
        for i, b in enumerate(self._tab_btns):
            b.setChecked(i == idx)

    def _apply_theme(self):
        c   = _api.theme.colors()
        g   = c.get("glow",           "#00c8ff")
        bg  = c.get("bg_dark",        "#0e1422")
        mid = c.get("bg_mid",         "#141c2e")
        tp  = c.get("text_primary",   "#dce8ff")
        ts  = c.get("text_secondary", "#8cbede")
        gr, gg, gb = int(g[1:3], 16), int(g[3:5], 16), int(g[5:7], 16)
        self.setStyleSheet(f"""
            QDialog    {{ background:{bg}; color:{tp}; }}
            #cs_hdr    {{ background:{mid}; border-bottom:1px solid rgba({gr},{gg},{gb},40); }}
            #cs_title_lbl {{ color:{g}; }}
            #cs_tbar   {{ background:{mid}; }}
            #cs_sep    {{ color:rgba({gr},{gg},{gb},32); }}
            QPushButton[objectName^="cs_tab_"] {{
                background:transparent; color:{ts};
                border:none; border-radius:6px;
                padding:2px 18px; font-size:11px;
                outline: none;
            }}
            QPushButton[objectName^="cs_tab_"]:hover   {{ background:rgba({gr},{gg},{gb},22); color:{tp}; }}
            QPushButton[objectName^="cs_tab_"]:checked  {{
                background:rgba({gr},{gg},{gb},34); color:{g};
                border-bottom:2px solid {g};
            }}
            QPushButton[objectName^="cs_tab_"]:focus   {{ outline: none; }}
            QStackedWidget {{ background:{bg}; }}
        """)
        for tab in (self.iva_tab, self.qa_tab, self.rs_tab):
            if tab:
                tab.apply_theme(c)


# ═══════════════════════════════════════════════════════════════════════════════
# Tab 1 — IVA Formatter
# ═══════════════════════════════════════════════════════════════════════════════

class IVAFormatterTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._src_path: Optional[str]  = None
        self._debounce: Optional[QTimer] = None
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(20, 18, 20, 14)
        outer.setSpacing(12)

        # ── openpyxl warning ─────────────────────────────────────────────────
        self._no_openpyxl_bar = QFrame()
        self._no_openpyxl_bar.setObjectName("iva_warn")
        wlay = QHBoxLayout(self._no_openpyxl_bar)
        wlay.setContentsMargins(10, 6, 10, 6)
        wlay.addWidget(QLabel("⚠  openpyxl is not installed — IVA formatting unavailable."))
        install_btn = QPushButton("Install openpyxl")
        install_btn.setObjectName("iva_install_btn")
        install_btn.setFixedWidth(150)
        install_btn.clicked.connect(_install_openpyxl)
        wlay.addStretch()
        wlay.addWidget(install_btn)
        self._no_openpyxl_bar.setVisible(not _OPENPYXL_OK)
        outer.addWidget(self._no_openpyxl_bar)

        # ── section: Manual Format ────────────────────────────────────────────
        manual_lbl = QLabel("MANUAL FORMAT")
        manual_lbl.setObjectName("iva_section_lbl")
        manual_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        outer.addWidget(manual_lbl)

        pick_row = QHBoxLayout()
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("No file selected — click Browse…")
        self._file_edit.setReadOnly(True)
        self._file_edit.setObjectName("iva_file_edit")
        browse_btn = QPushButton("Browse…")
        browse_btn.setObjectName("iva_btn")
        browse_btn.setFixedWidth(88)
        browse_btn.clicked.connect(self._browse)
        pick_row.addWidget(self._file_edit)
        pick_row.addWidget(browse_btn)
        outer.addLayout(pick_row)

        # preview
        self._preview_lbl = QLabel("New filename will appear here after selecting a file.")
        self._preview_lbl.setObjectName("iva_preview_lbl")
        self._preview_lbl.setWordWrap(True)
        outer.addWidget(self._preview_lbl)

        # action row
        action_row = QHBoxLayout()
        action_row.setSpacing(10)
        self._format_btn = QPushButton("Format & Save")
        self._format_btn.setObjectName("iva_primary_btn")
        self._format_btn.setFixedHeight(34)
        self._format_btn.setEnabled(False)
        self._format_btn.clicked.connect(self._format_and_save)
        self._open_dest_btn = QPushButton("Open Destination Folder")
        self._open_dest_btn.setObjectName("iva_btn")
        self._open_dest_btn.setEnabled(False)
        self._open_dest_btn.clicked.connect(self._open_dest)
        action_row.addWidget(self._format_btn)
        action_row.addWidget(self._open_dest_btn)
        action_row.addStretch()
        outer.addLayout(action_row)

        outer.addWidget(_sep())

        # ── section: Output Directory ─────────────────────────────────────────
        out_lbl = QLabel("OUTPUT DIRECTORY")
        out_lbl.setObjectName("iva_section_lbl")
        out_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        outer.addWidget(out_lbl)

        out_row = QHBoxLayout()
        out_row.setSpacing(6)
        self._out_edit = QLineEdit()
        self._out_edit.setObjectName("iva_file_edit")
        self._out_edit.setPlaceholderText(str(_DEFAULT_OUTPUT_DIR))
        self._out_edit.setReadOnly(True)
        saved_out = _api.settings.value("output_dir", "") if _api else ""
        self._out_edit.setText(saved_out or str(_DEFAULT_OUTPUT_DIR))
        out_browse_btn = QPushButton("Change…")
        out_browse_btn.setObjectName("iva_btn")
        out_browse_btn.setFixedWidth(80)
        out_browse_btn.clicked.connect(self._browse_output_dir)
        out_reset_btn = QPushButton("Reset")
        out_reset_btn.setObjectName("iva_btn")
        out_reset_btn.setFixedWidth(56)
        out_reset_btn.clicked.connect(self._reset_output_dir)
        out_open_btn = QPushButton("Open")
        out_open_btn.setObjectName("iva_btn")
        out_open_btn.setFixedWidth(56)
        out_open_btn.clicked.connect(self._open_dest)
        out_row.addWidget(self._out_edit)
        out_row.addWidget(out_browse_btn)
        out_row.addWidget(out_reset_btn)
        out_row.addWidget(out_open_btn)
        outer.addLayout(out_row)

        outer.addWidget(_sep())

        # ── section: Auto Format from Downloads ───────────────────────────────
        auto_lbl = QLabel("AUTO FORMAT FROM DOWNLOADS")
        auto_lbl.setObjectName("iva_section_lbl")
        auto_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        outer.addWidget(auto_lbl)

        auto_row = QHBoxLayout()
        self._auto_chk = QCheckBox("Enable — watch Downloads folder for new IVA files")
        self._auto_chk.setObjectName("iva_chk")
        self._auto_chk.setChecked(_watcher_enabled)
        self._auto_chk.toggled.connect(self._toggle_watcher)
        auto_row.addWidget(self._auto_chk)
        auto_row.addStretch()
        outer.addLayout(auto_row)

        info_lbl = QLabel(
            f"Watching: {_DOWNLOADS}\n"
            "Pattern: CLIENT_IVA_YYYYMMDDHHMI Category.xlsx  "
            "(e.g. SBBISMARCK_IVA_202601020001 SSL and TLS.xlsx)"
        )
        info_lbl.setObjectName("iva_info_lbl")
        info_lbl.setWordWrap(True)
        outer.addWidget(info_lbl)

        outer.addWidget(_sep())

        # ── section: Log ──────────────────────────────────────────────────────
        log_header = QHBoxLayout()
        log_lbl = QLabel("LOG")
        log_lbl.setObjectName("iva_section_lbl")
        log_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        clear_log_btn = QPushButton("Clear")
        clear_log_btn.setObjectName("iva_btn")
        clear_log_btn.setFixedSize(56, 22)
        clear_log_btn.clicked.connect(lambda: self._log.clear())
        log_header.addWidget(log_lbl)
        log_header.addStretch()
        log_header.addWidget(clear_log_btn)
        outer.addLayout(log_header)

        self._log = QTextEdit()
        self._log.setObjectName("iva_log")
        self._log.setReadOnly(True)
        self._log.setFont(QFont("Consolas", 9))
        self._log.setMinimumHeight(90)
        self._log.document().setMaximumBlockCount(500)  # prevent unbounded growth
        outer.addWidget(self._log)

    # ── actions ───────────────────────────────────────────────────────────────
    def _browse(self):
        try:
            path = _api.files.open_dialog(
                "Select IVA Spreadsheet", "Excel Files (*.xlsx);;All Files (*)")
            if not path:
                return
            self._src_path = path
            self._file_edit.setText(path)
            self._update_preview(path)
        except Exception as e:
            _api.log(f"CySec browse error: {e}")
            _api.toast("Error opening file dialog.", "error")

    def _update_preview(self, path: str):
        if not _OPENPYXL_OK:
            self._preview_lbl.setText("⚠  openpyxl required — install it first.")
            self._format_btn.setEnabled(False)
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                sheet_name = wb.active.title
            finally:
                wb.close()
            clean = _clean_stem(Path(path).name)
            preview = f"📄  New filename:  {clean} {sheet_name} [COMPLETE|WIP].xlsx"
            self._preview_lbl.setText(preview)
            self._format_btn.setEnabled(True)
            self._open_dest_btn.setEnabled(True)
        except Exception as e:
            self._preview_lbl.setText(f"⚠  Could not read workbook: {e}")
            self._format_btn.setEnabled(False)

    def _format_and_save(self):
        if not self._src_path:
            return
        try:
            src = Path(self._src_path)
            self.append_log(f"Formatting: {src.name} …")
            result = _format_iva_file(src)
            if result:
                new_name, status = result
                self.append_log(f"✔  Saved: {new_name}  [{status}]")
                self._preview_lbl.setText(f"✔  Saved: {src.parent / new_name}")
                _api.toast(f"IVA formatted → {new_name}  [{status}]", "success")
        except Exception as e:
            self.append_log(f"✘  Error: {e}")
            _api.toast(f"Formatting failed: {e}", "error")
            _api.log(f"CySec format error: {e}")

    def _open_dest(self):
        dest = str(_get_output_dir())
        try:
            subprocess.Popen(["explorer", dest])
        except Exception as e:
            _api.log(f"CySec open dest error: {e}")

    def _browse_output_dir(self):
        try:
            folder = _api.files.open_folder_dialog("Select Output Directory")
            if not folder:
                return
            p = Path(folder)
            if p.resolve() == _DOWNLOADS.resolve():
                _api.toast("Output directory cannot be the Downloads folder.", "error")
                return
            _api.settings.set("output_dir", str(p))
            self._out_edit.setText(str(p))
            self.append_log(f"Output directory set to: {p}")
        except Exception as e:
            _api.log(f"CySec browse output dir error: {e}")
            _api.toast("Error selecting output directory.", "error")

    def _reset_output_dir(self):
        _api.settings.remove("output_dir")
        self._out_edit.setText(str(_DEFAULT_OUTPUT_DIR))
        self.append_log(f"Output directory reset to default: {_DEFAULT_OUTPUT_DIR}")

    def _toggle_watcher(self, checked: bool):
        try:
            _set_watcher(checked)
            state = "enabled" if checked else "disabled"
            self.append_log(f"Auto Format from Downloads {state}.")
            _api.toast(f"Auto Format {state}.", "info")
        except Exception as e:
            _api.log(f"CySec toggle watcher error: {e}")

    def append_log(self, msg: str):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self._log.append(f"[{ts}]  {msg}")

    # ── theme ─────────────────────────────────────────────────────────────────
    def apply_theme(self, c: dict):
        g   = c.get("glow",           "#00c8ff")
        bg  = c.get("bg_dark",        "#0e1422")
        mid = c.get("bg_mid",         "#141c2e")
        tp  = c.get("text_primary",   "#dce8ff")
        ts  = c.get("text_secondary", "#8cbede")
        td  = c.get("text_dim",       "#50788f")
        ab  = c.get("accent_amber",   "#ffaa14")
        ar  = c.get("accent_red",     "#e63c46")
        gr,  gg,  gb  = int(g[1:3],  16), int(g[3:5],  16), int(g[5:7],  16)
        abr, abg, abb = int(ab[1:3], 16), int(ab[3:5], 16), int(ab[5:7], 16)
        arr, arg, arb = int(ar[1:3], 16), int(ar[3:5], 16), int(ar[5:7], 16)
        tdr, tdg, tdb = int(td[1:3], 16), int(td[3:5], 16), int(td[5:7], 16)
        self.setStyleSheet(f"""
            QWidget       {{ background:{bg}; color:{tp}; }}
            #iva_warn     {{ background:rgba({arr},{arg},{arb},40); border:1px solid rgba({arr},{arg},{arb},96);
                             border-radius:6px; color:{ar}; }}
            #iva_section_lbl {{ color:{g}; letter-spacing:1px; }}
            #iva_preview_lbl {{ color:{ts}; font-style:italic; }}
            #iva_info_lbl    {{ color:{td}; font-size:10px; }}
            #iva_file_edit   {{ background:{mid}; color:{tp}; border:1px solid rgba({gr},{gg},{gb},48);
                                border-radius:4px; padding:4px 8px;
                                selection-background-color:rgba({gr},{gg},{gb},80); selection-color:{tp}; }}
            #iva_file_edit:focus {{ border-color:rgba({gr},{gg},{gb},112); outline: none; }}
            #iva_chk         {{ color:{tp}; spacing:8px; outline: none; }}
            #iva_chk::indicator         {{ width:16px; height:16px;
                                           border:1px solid rgba({gr},{gg},{gb},80); border-radius:3px;
                                           background:{mid}; }}
            #iva_chk::indicator:checked {{ background:{g}; border-color:{g}; }}
            #iva_log         {{ background:{mid}; color:{ts}; border:1px solid rgba({gr},{gg},{gb},32);
                                border-radius:6px; padding:4px;
                                selection-background-color:rgba({gr},{gg},{gb},80); selection-color:{tp}; }}
            #iva_btn {{ background:{mid}; color:{tp}; border:1px solid rgba({gr},{gg},{gb},48);
                        border-radius:4px; padding:3px 10px; outline: none; }}
            #iva_btn:hover   {{ border-color:rgba({gr},{gg},{gb},112); color:{g}; }}
            #iva_btn:focus   {{ outline: none; border-color:rgba({gr},{gg},{gb},96); }}
            #iva_install_btn {{ background:rgba({abr},{abg},{abb},32); color:{ab}; border:1px solid rgba({abr},{abg},{abb},96);
                                border-radius:4px; padding:3px 10px; outline: none; }}
            #iva_install_btn:hover {{ background:rgba({abr},{abg},{abb},64); }}
            #iva_install_btn:focus {{ outline: none; }}
            #iva_primary_btn {{ background:rgba({gr},{gg},{gb},34); color:{g}; border:1px solid rgba({gr},{gg},{gb},96);
                                 border-radius:6px; font-weight:bold; padding:4px 20px;
                                 outline: none; }}
            #iva_primary_btn:hover    {{ background:rgba({gr},{gg},{gb},56); }}
            #iva_primary_btn:focus    {{ outline: none; }}
            #iva_primary_btn:disabled {{ color:{td}; border-color:rgba({tdr},{tdg},{tdb},48);
                                         background:transparent; }}
            #cs_sep {{ color:rgba({gr},{gg},{gb},32); }}
        """)


# ═══════════════════════════════════════════════════════════════════════════════
# Tab 2 — Quick Access
# ═══════════════════════════════════════════════════════════════════════════════

class QuickAccessTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(24, 24, 24, 24)
        outer.setSpacing(14)

        hdr = QLabel("Quick File Access")
        hdr.setObjectName("qa_hdr")
        hdr.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        outer.addWidget(hdr)

        sub = QLabel("Open frequently used directories instantly.")
        sub.setObjectName("qa_sub")
        outer.addWidget(sub)

        outer.addSpacing(10)

        cards = QHBoxLayout()
        cards.setSpacing(18)

        self._scripts_btn = self._card(
            "📁", "Scripts Folder", _SCRIPTS_PATH, self._open_scripts)
        self._forti_btn = self._card(
            "🛡", "Fortinet FortiEMS", _FORTI_PATH, self._open_forti)

        cards.addWidget(self._scripts_btn)
        cards.addWidget(self._forti_btn)
        cards.addStretch()
        outer.addLayout(cards)
        outer.addStretch()

    def _card(self, icon: str, title: str, path: str, fn) -> QFrame:
        card = QFrame()
        card.setObjectName("qa_card")
        card.setMinimumSize(210, 150)
        card.setMaximumSize(280, 200)
        card.setCursor(Qt.CursorShape.PointingHandCursor)
        lay = QVBoxLayout(card)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.setSpacing(6)

        ico = QLabel(icon)
        ico.setFont(QFont("Segoe UI Emoji", 26))
        ico.setAlignment(Qt.AlignmentFlag.AlignCenter)

        t_lbl = QLabel(title)
        t_lbl.setObjectName("qa_card_title")
        t_lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        t_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)

        p_lbl = QLabel(path)
        p_lbl.setObjectName("qa_card_path")
        p_lbl.setFont(QFont("Consolas", 8))
        p_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        p_lbl.setWordWrap(True)

        lay.addWidget(ico)
        lay.addWidget(t_lbl)
        lay.addWidget(p_lbl)

        # Click via mousePressEvent override
        card._fn = fn
        card.mousePressEvent = lambda event, _fn=fn: _fn()
        return card

    def _open_scripts(self):
        try:
            subprocess.Popen(["explorer", _SCRIPTS_PATH])
        except Exception as e:
            _api.log(f"CySec open scripts error: {e}")
            _api.toast(f"Could not open {_SCRIPTS_PATH}", "error")

    def _open_forti(self):
        try:
            subprocess.Popen(["explorer", _FORTI_PATH])
        except Exception as e:
            _api.log(f"CySec open forti error: {e}")
            _api.toast(f"Could not open {_FORTI_PATH}", "error")

    def apply_theme(self, c: dict):
        g   = c.get("glow",           "#00c8ff")
        bg  = c.get("bg_dark",        "#0e1422")
        mid = c.get("bg_mid",         "#141c2e")
        tp  = c.get("text_primary",   "#dce8ff")
        ts  = c.get("text_secondary", "#8cbede")
        td  = c.get("text_dim",       "#50788f")
        gr, gg, gb = int(g[1:3], 16), int(g[3:5], 16), int(g[5:7], 16)
        self.setStyleSheet(f"""
            QWidget         {{ background:{bg}; color:{tp}; }}
            #qa_hdr         {{ color:{g}; }}
            #qa_sub         {{ color:{ts}; }}
            #qa_card        {{ background:{mid}; border:1px solid rgba({gr},{gg},{gb},48);
                               border-radius:14px; }}
            #qa_card:hover  {{ background:rgba({gr},{gg},{gb},24); border-color:rgba({gr},{gg},{gb},112); }}
            #qa_card_title  {{ color:{tp}; }}
            #qa_card_path   {{ color:{td}; }}
        """)


# ═══════════════════════════════════════════════════════════════════════════════
# Tab 3 — Response Search
# ═══════════════════════════════════════════════════════════════════════════════

class ResponseSearchTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._active_cat: Optional[str] = None
        self._filtered: list[dict]      = list(_RESPONSES)
        # Single persistent debounce timer — restarted on each keystroke.
        # Creating a new QTimer per keystroke leaks objects.
        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.timeout.connect(self._refresh_results)
        self._cat_btns: dict[str, QPushButton] = {}
        self._build_ui()
        self._refresh_results()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 12, 14, 10)
        outer.setSpacing(8)

        # ── search row ────────────────────────────────────────────────────────
        search_row = QHBoxLayout()
        search_row.setSpacing(8)
        self._search_edit = QLineEdit()
        self._search_edit.setObjectName("rs_search")
        self._search_edit.setPlaceholderText("🔍  Search by response text or tag…")
        self._search_edit.setFixedHeight(34)
        self._search_edit.textChanged.connect(self._on_search_changed)
        self._clear_btn = QPushButton("✕")
        self._clear_btn.setObjectName("rs_clear_btn")
        self._clear_btn.setFixedSize(34, 34)
        self._clear_btn.setToolTip("Clear search")
        self._clear_btn.clicked.connect(self._search_edit.clear)
        search_row.addWidget(self._search_edit)
        search_row.addWidget(self._clear_btn)
        outer.addLayout(search_row)

        # ── category filter chips ────────────────────────────────────────────
        chip_scroll = QScrollArea()
        chip_scroll.setObjectName("rs_chip_scroll")
        chip_scroll.setFixedHeight(46)
        chip_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        chip_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        chip_scroll.setWidgetResizable(True)
        chip_scroll.setFrameShape(QFrame.Shape.NoFrame)

        chips_w = QWidget()
        chips_w.setObjectName("rs_chips_w")
        chips_lay = QHBoxLayout(chips_w)
        chips_lay.setContentsMargins(2, 4, 2, 4)
        chips_lay.setSpacing(6)

        # "All" chip
        all_btn = QPushButton("All")
        all_btn.setObjectName("rs_chip_all")
        all_btn.setCheckable(True)
        all_btn.setChecked(True)
        all_btn.setFixedHeight(28)
        all_btn.clicked.connect(lambda: self._set_category(None))
        chips_lay.addWidget(all_btn)
        self._cat_btns["__all__"] = all_btn

        for cat in _ALL_CATEGORIES:
            short = (cat[:22] + "…") if len(cat) > 24 else cat
            btn = QPushButton(short)
            btn.setObjectName("rs_chip")
            btn.setCheckable(True)
            btn.setChecked(False)
            btn.setFixedHeight(28)
            btn.setToolTip(cat)
            btn.clicked.connect(lambda _ch, _c=cat: self._set_category(_c))
            chips_lay.addWidget(btn)
            self._cat_btns[cat] = btn

        chips_lay.addStretch()
        chip_scroll.setWidget(chips_w)
        outer.addWidget(chip_scroll)

        # ── splitter: results | preview ───────────────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setObjectName("rs_splitter")
        splitter.setHandleWidth(6)

        # Results list
        list_w = QWidget()
        list_w.setObjectName("rs_list_pane")
        list_lay = QVBoxLayout(list_w)
        list_lay.setContentsMargins(0, 0, 0, 0)
        list_lay.setSpacing(4)

        self._count_lbl = QLabel(f"Results: {len(_RESPONSES)}")
        self._count_lbl.setObjectName("rs_count_lbl")
        self._count_lbl.setFont(QFont("Segoe UI", 8))
        list_lay.addWidget(self._count_lbl)

        self._list = QListWidget()
        self._list.setObjectName("rs_list")
        self._list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._list.currentItemChanged.connect(self._on_select)
        list_lay.addWidget(self._list)

        splitter.addWidget(list_w)

        # Preview pane
        prev_w = QWidget()
        prev_w.setObjectName("rs_prev_pane")
        prev_lay = QVBoxLayout(prev_w)
        prev_lay.setContentsMargins(10, 0, 0, 0)
        prev_lay.setSpacing(6)

        self._prev_cat  = QLabel("")
        self._prev_cat.setObjectName("rs_prev_cat")
        self._prev_cat.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        self._prev_cat.setWordWrap(True)

        self._prev_text = QTextEdit()
        self._prev_text.setObjectName("rs_prev_text")
        self._prev_text.setReadOnly(True)
        self._prev_text.setFont(QFont("Segoe UI", 10))

        self._prev_tags = QLabel("")
        self._prev_tags.setObjectName("rs_prev_tags")
        self._prev_tags.setFont(QFont("Segoe UI", 8))
        self._prev_tags.setWordWrap(True)

        copy_row = QHBoxLayout()
        self._copy_btn = QPushButton("📋  Copy Response")
        self._copy_btn.setObjectName("rs_copy_btn")
        self._copy_btn.setEnabled(False)
        self._copy_btn.clicked.connect(self._copy_response)
        copy_row.addWidget(self._copy_btn)
        copy_row.addStretch()

        prev_lay.addWidget(self._prev_cat)
        prev_lay.addWidget(self._prev_text, stretch=1)
        prev_lay.addWidget(self._prev_tags)
        prev_lay.addLayout(copy_row)

        splitter.addWidget(prev_w)
        splitter.setSizes([300, 620])

        outer.addWidget(splitter, stretch=1)

        # ── bottom action bar ─────────────────────────────────────────────────
        bot = QHBoxLayout()
        full_btn = QPushButton("📄  Open Full Response Sheet")
        full_btn.setObjectName("rs_full_btn")
        full_btn.clicked.connect(self._open_full_sheet)
        self._total_lbl = QLabel(f"{len(_RESPONSES)} total responses")
        self._total_lbl.setObjectName("rs_total_lbl")
        self._total_lbl.setFont(QFont("Segoe UI", 8))
        bot.addWidget(full_btn)
        bot.addStretch()
        bot.addWidget(self._total_lbl)
        outer.addLayout(bot)

    # ── logic ─────────────────────────────────────────────────────────────────
    def _on_search_changed(self, _text: str):
        self._debounce.start(200)

    def _set_category(self, cat: Optional[str]):
        self._active_cat = cat
        for k, btn in self._cat_btns.items():
            if cat is None:
                btn.setChecked(k == "__all__")
            else:
                btn.setChecked(k == cat)
        self._refresh_results()

    def _refresh_results(self):
        query = self._search_edit.text().strip().lower()
        results: list[dict] = []
        for r in _RESPONSES:
            if self._active_cat and r["category"] != self._active_cat:
                continue
            if query:
                if query not in r["text"].lower() and not any(query in t for t in r["tags"]):
                    continue
            results.append(r)
        self._filtered = results
        self._count_lbl.setText(f"Results: {len(results)}")
        self._list.clear()
        for r in results:
            snippet = r["text"][:82].replace("\n", " ")
            if len(r["text"]) > 82:
                snippet += "…"
            item = QListWidgetItem()
            item.setText(f"{r['category']}\n{snippet}")
            item.setData(Qt.ItemDataRole.UserRole, r["id"])
            self._list.addItem(item)
        # Clear preview if nothing selected
        if not results:
            self._prev_cat.setText("")
            self._prev_text.clear()
            self._prev_tags.setText("")
            self._copy_btn.setEnabled(False)

    def _on_select(self, current: Optional[QListWidgetItem], _prev):
        if current is None:
            return
        rid = current.data(Qt.ItemDataRole.UserRole)
        match = next((r for r in _RESPONSES if r["id"] == rid), None)
        if not match:
            return
        self._prev_cat.setText(f"📂  {match['category']}")
        self._prev_text.setPlainText(match["text"])
        tag_str = "  ".join(f"#{t}" for t in match["tags"])
        self._prev_tags.setText(tag_str)
        self._copy_btn.setEnabled(True)

    def _copy_response(self):
        try:
            text = self._prev_text.toPlainText()
            if text:
                _api.clipboard.set_text(text)
                _api.toast("Response copied to clipboard.", "success")
        except Exception as e:
            _api.log(f"CySec copy error: {e}")

    def _open_full_sheet(self):
        try:
            dest = _api.plugin_dir / "responses_full.txt"
            lines: list[str] = []
            current_cat = ""
            for r in _RESPONSES:
                if r["category"] != current_cat:
                    current_cat = r["category"]
                    lines.append(f"\n{'═' * 70}")
                    lines.append(f"  {current_cat.upper()}")
                    lines.append(f"{'═' * 70}\n")
                lines.append(r["text"])
                lines.append("")
            _api.files.write_text(dest, "\n".join(lines))
            _api.open_url(str(dest))
        except Exception as e:
            _api.log(f"CySec open_full_sheet error: {e}")
            _api.toast("Could not open full response sheet.", "error")

    # ── theme ─────────────────────────────────────────────────────────────────
    def apply_theme(self, c: dict):
        g   = c.get("glow",           "#00c8ff")
        bg  = c.get("bg_dark",        "#0e1422")
        mid = c.get("bg_mid",         "#141c2e")
        tp  = c.get("text_primary",   "#dce8ff")
        ts  = c.get("text_secondary", "#8cbede")
        td  = c.get("text_dim",       "#50788f")
        ab  = c.get("accent_blue",    "#1e96ff")
        at  = c.get("accent_teal",    "#00d2be")
        gr,  gg,  gb  = int(g[1:3],  16), int(g[3:5],  16), int(g[5:7],  16)
        abr, abg, abb = int(ab[1:3], 16), int(ab[3:5], 16), int(ab[5:7], 16)
        tdr, tdg, tdb = int(td[1:3], 16), int(td[3:5], 16), int(td[5:7], 16)
        self.setStyleSheet(f"""
            QWidget          {{ background:{bg}; color:{tp}; }}
            #rs_search       {{ background:{mid}; color:{tp};
                                border:1px solid rgba({gr},{gg},{gb},64); border-radius:6px;
                                padding:2px 10px; font-size:12px;
                                selection-background-color:rgba({gr},{gg},{gb},80); selection-color:{tp}; }}
            #rs_search:focus {{ border-color:{g}; outline: none; }}
            #rs_clear_btn    {{ background:{mid}; color:{td};
                                border:1px solid rgba({gr},{gg},{gb},37); border-radius:6px;
                                font-size:13px; font-weight:bold; outline: none; }}
            #rs_clear_btn:hover {{ color:{tp}; border-color:rgba({gr},{gg},{gb},96); }}
            #rs_clear_btn:focus {{ outline: none; border-color:rgba({gr},{gg},{gb},80); }}
            #rs_chips_w      {{ background:{bg}; }}
            #rs_chip_scroll  {{ background:{bg}; border:none; }}
            #rs_chip_all, #rs_chip {{
                background:{mid}; color:{ts};
                border:1px solid rgba({gr},{gg},{gb},40); border-radius:12px;
                padding:2px 12px; font-size:10px; outline: none;
            }}
            #rs_chip_all:hover, #rs_chip:hover {{
                background:rgba({gr},{gg},{gb},22); color:{tp}; border-color:rgba({gr},{gg},{gb},96);
            }}
            #rs_chip_all:checked, #rs_chip:checked {{
                background:rgba({gr},{gg},{gb},37); color:{g}; border-color:{g};
            }}
            #rs_chip_all:focus, #rs_chip:focus {{ outline: none; }}
            #rs_list_pane    {{ background:{bg}; }}
            #rs_count_lbl    {{ color:{td}; padding:0 2px; }}
            #rs_list         {{ background:{mid}; color:{tp};
                                border:1px solid rgba({gr},{gg},{gb},32); border-radius:6px;
                                outline:none; }}
            #rs_list::item   {{ padding:6px 8px; border-bottom:1px solid rgba({gr},{gg},{gb},16);
                                font-size:10px; color:{ts}; }}
            #rs_list::item:selected {{
                background:rgba({gr},{gg},{gb},34); color:{tp}; border-radius:4px;
            }}
            #rs_list::item:hover {{ background:rgba({gr},{gg},{gb},16); }}
            #rs_prev_pane    {{ background:{bg}; }}
            #rs_prev_cat     {{ color:{at}; padding:2px 0; }}
            #rs_prev_text    {{ background:{mid}; color:{tp};
                                border:1px solid rgba({gr},{gg},{gb},32); border-radius:6px;
                                padding:8px; line-height:1.5;
                                selection-background-color:rgba({gr},{gg},{gb},80); selection-color:{tp}; }}
            #rs_prev_tags    {{ color:{td}; font-size:9px; padding-top:2px; }}
            #rs_copy_btn     {{ background:rgba({abr},{abg},{abb},34); color:{ab};
                                border:1px solid rgba({abr},{abg},{abb},85); border-radius:6px;
                                padding:4px 14px; font-weight:bold; outline: none; }}
            #rs_copy_btn:hover    {{ background:rgba({abr},{abg},{abb},64); }}
            #rs_copy_btn:focus    {{ outline: none; }}
            #rs_copy_btn:disabled {{ color:{td}; border-color:rgba({tdr},{tdg},{tdb},48);
                                     background:transparent; }}
            #rs_full_btn     {{ background:{mid}; color:{ts};
                                border:1px solid rgba({gr},{gg},{gb},48); border-radius:6px;
                                padding:4px 14px; outline: none; }}
            #rs_full_btn:hover {{ border-color:rgba({gr},{gg},{gb},112); color:{g}; }}
            #rs_full_btn:focus {{ outline: none; border-color:rgba({gr},{gg},{gb},96); }}
            #rs_total_lbl    {{ color:{td}; }}
            QSplitter::handle {{ background:rgba({gr},{gg},{gb},24); }}
            QScrollBar:horizontal {{
                background:{mid}; height:6px; border-radius:3px;
            }}
            QScrollBar::handle:horizontal {{
                background:rgba({gr},{gg},{gb},80); border-radius:3px; min-width:20px;
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width:0;
            }}
        """)
